import bs4
import numpy
import requests
from sense import *
from math import exp
from numpy import mean
from openpyxl import Workbook, load_workbook
numpy.seterr(all='raise')
import scipy.stats as stats

class Word:

    def __init__(self, sheet, title, part_of_speech="all"):
        self.word = title[:-5:]
        self.PoS = part_of_speech
        self.senses = self.make_sense_list(sheet, part_of_speech)
        self.length = len(self.senses)
        self.viable = True  #becomes false when the word has correlations that are not viable, or has no non-baseline senses
        if len(self.senses) <= 1: #or self.senses[0].date == self.senses[-1].date:
            self.bl_senses = self.senses
            self.nb_senses = []
            self.start_date = self.senses[0].date
        else:
            index = 0
            while True:
                if index >= self.length:
                    break
                if self.senses[index].date != self.senses[index+1].date:
                    break
                index += 1
            self.bl_senses = self.senses[:index+1:]
            self.nb_senses = self.senses[index+1::]

        if self.nb_senses:
            self.num_nb_senses = len(self.nb_senses)
            self.num_bl_senses = self.length - self.num_nb_senses
            self.num_senses = self.num_nb_senses + self.num_bl_senses
            self.num_exist_senses = len([s for s in self.senses if s.end_date == 3000])
            self.nb_dates = [nb_sense.date for nb_sense in self.nb_senses]
            self.vs_baseline = [Baseline(bl_sense, self.nb_senses) for bl_sense in self.bl_senses]  #list of baseline senses
            self.vs_baseline = [baseline for baseline in self.vs_baseline if baseline.pearsonR]

    def make_sense_list(self, sheet, part_of_speech):
        """Takes a word and searches for it within the HTE database,
        returning a list with the same information in a sense
        object with, ordered a date, word_form, identifiers, and
        categories"""

        ws = sheet.active

        if part_of_speech == "a":
            part_of_speech = "aj"
        if part_of_speech == "adv":
            part_of_speech = "av"

        sense_list = []

        r = 1

        while ws.cell(row=r, column=1).value:
            date = ws.cell(row=r, column=1).value
            word_form = ws.cell(row=r, column=2).value
            identifiers = ws.cell(row=r, column=3).value
            categories = ws.cell(row=r, column=4).value
            PoS = ws.cell(row=r, column=5).value
            r += 1
            s = Sense(self.word, date, word_form, identifiers, categories, PoS)
            if s.date:
                sense_list.append(s)
        s_oe = [s for s in sense_list if s.date == "OE"]
        s_non_oe = [s for s in sense_list if s.date != "OE"]
        return s_oe + sorted(s_non_oe, key=lambda x : x.date)

    def __repr__(self):
        return self.label

    def __str__(self):
        return self.label

class Baseline:

    def __init__(self, bl_sense, nb_senses):
        self.bl_sense = bl_sense
        self.nb_senses = nb_senses
        self.nb_dates = [nb_sense.date for nb_sense in nb_senses]
        self.sim_scores = [self.calculate_score(bl_sense, nb_sense) for nb_sense in nb_senses]
        try:
            self.pearsonR = numpy.corrcoef(self.nb_dates, self.sim_scores).tolist()[0][1]
        except:
            self.pearsonR = False

    def calculate_score(self, bl_sense, nb_sense):
        similarity_score = 0
        list_range = min(len(bl_sense.listed_cat), len(nb_sense.listed_cat))
        for r in range(list_range):
            if bl_sense.listed_cat[r] == nb_sense.listed_cat[r]:
                similarity_score += 1
            else:
                break
        return exp(-similarity_score)
