from word_data_structures import *
from openpyxl import Workbook, load_workbook
from scipy import stats
import random
import numpy as np
import matplotlib.pyplot as plt

list = load_workbook("_word_list.xlsx")
ws = list.active
i = 2
models = ["Exemplar", "Prototype", "Progenitor", "Local", "NN Chain"]

while ws.cell(row=i, column=1).value:
    str_word = ws.cell(row=i, column=1).value
    curr_xlsx = load_workbook(str_word + ".xlsx")
    word = Word(curr_xlsx, str_word + ".xlsx")
    print(i, flush=True)

    ws.cell(row=i, column=2).value = np.sum(np.log(create_score_comparisons(word, over_all_exemplar, True))) - np.sum(np.log(null_probabilities(word)))
    ws.cell(row=i, column=3).value = np.sum(np.log(create_score_comparisons(word, over_all_dynamic, True))) - np.sum(np.log(null_probabilities(word)))
    ws.cell(row=i, column=4).value = np.sum(np.log(create_score_comparisons(word, over_all_prototype, False))) - np.sum(np.log(null_probabilities(word)))
    ws.cell(row=i, column=5).value = np.sum(np.log(create_score_comparisons(word, over_all_simple, False))) - np.sum(np.log(null_probabilities(word)))
    ws.cell(row=i, column=6).value = np.sum(np.log(create_score_comparisons(word, over_all_complex, True))) - np.sum(np.log(null_probabilities(word)))


    scores = np.array([ws.cell(row=i, column=col).value for col in range(2, 7)])
    max_v = scores.max()
    min_indices = [i for i, v in enumerate(scores) if v == max_v]
    if len(min_indices) > 1:
        print("Tied")
        ws.cell(row=i, column=7).value = "Tied"
    else:
        print(models[min_indices[0]])
        ws.cell(row=i, column=7).value = models[min_indices[0]]
    i += 1
    if i % 10 == 0:
        list.save("_word_list.xlsx")
list.save("_word_list.xlsx")


