from word_models import *
from openpyxl import Workbook, load_workbook
import numpy as np

def list_iteration(c, end, start=2, PoS=False):
    """Perform the analysis on corpus c (an Excel file ending with '.xlsx.,
    supplied with the starting and ending table rows from the Excel file. PoS
    is initially set to False, which will consider all senses of a word, and
    will only take all senses that fall into a specific part-of-speech otherwise."""

    corpus = load_workbook(c)
    ws = corpus.active

    labels = ["Progenitor", "Markov", "Exemplar", "Chaining", "Dynamic"]
    for i in range(2, end+1):
        try:
            if PoS:
                word = Word(ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value)
            else:
                word = Word(ws.cell(row=i, column=1).value)
            print(str(i) + " | " + str(word.label), flush=True)

            ws.cell(row=i, column=3).value = np.sum(np.log(create_score_comparisons(word, over_all_progenitor, False))) - np.sum(np.log(null_probabilities(word)))
            ws.cell(row=i, column=4).value = np.sum(np.log(create_score_comparisons(word, over_all_markov, False))) - np.sum(np.log(null_probabilities(word)))
            ws.cell(row=i, column=5).value = np.sum(np.log(create_score_comparisons(word, over_all_exemplar, True))) - np.sum(np.log(null_probabilities(word)))
            ws.cell(row=i, column=6).value = np.sum(np.log(create_score_comparisons(word, over_all_chaining, True))) - np.sum(np.log(null_probabilities(word)))
            ws.cell(row=i, column=7).value = np.sum(np.log(create_score_comparisons(word, over_all_dynamic, True))) - np.sum(np.log(null_probabilities(word)))
            ws.cell(row=i, column=9).value = word.num_senses
            ws.cell(row=i, column=10).value = word.num_exist_senses
            ws.cell(row=i, column=11).value = word.bl_senses[0].date

            scores = np.array([ws.cell(row=i, column=col).value for col in range(3, 8)])
            max_v = scores.max()
            min_indices = [i for i, v in enumerate(scores) if v == max_v]
            if len(min_indices) > 1:
                print("Tied")
                ws.cell(row=i, column=8).value = "Tied"
            else:
                print(labels[min_indices[0]])
                ws.cell(row=i, column=8).value = labels[min_indices[0]]
        except:
            ws.cell(row=i, column=3).value = "FAIL"
            print("FAIL", flush=True)
        if i % 20 == 0:
            corpus.save(c)
    corpus.save(c)

list_iteration("BNC Models (Top 500).xlsx", 502, PoS=True)