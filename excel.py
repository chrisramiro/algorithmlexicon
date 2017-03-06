from openpyxl import Workbook
import os

def write_sense_list(word):

    wb = Workbook()
    ws1 = wb.active

    ws1['A1'], ws1['B1'], ws1['C1'], ws1['D1'], ws1['E1'], ws1['F1'] = "Start Date", "End Date", "Word Form", "Identifiers", "Categories", "PoS"
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 110
    ws1.column_dimensions['E'].width = 35
    for num in range(len(word.senses)):
        ws1.cell(row=num+2, column=1).value = word.senses[num].date
        if word.senses[num].end_date != 3000:
            ws1.cell(row=num+2, column=2).value = word.senses[num].end_date
        ws1.cell(row=num+2, column=3).value = word.senses[num].word_form
        ws1.cell(row=num+2, column=4).value = word.senses[num].identifiers
        ws1.cell(row=num+2, column=5).value = word.senses[num].categories
        ws1.cell(row=num+2, column=6).value = word.senses[num].PoS

    save_loc =  os.getcwd() + "\\" + "[" + word.PoS + "] " + word.label + ".xlsx"
    wb.save(save_loc)
    return save_loc