
from openpyxl import Workbook, load_workbook
import scipy.stats as stats

def do_print(wb, end):
    sim = load_workbook(wb)
    ws = sim.active
    score = []
    f, d, dr = [], [], []
    maxf = []

    for i in range(2, end):
        if ws.cell(row=i, column=5).value != ":(" and ws.cell(row=i, column=5).value:
            score.append(ws.cell(row=i, column=4).value)
            d.append(ws.cell(row=i, column=6).value)
            maxf.append(ws.cell(row=i, column=7).value)

    print(wb)
    print("d max: " + str(stats.spearmanr(score, d)))
    print("f max " + str(stats.spearmanr(score, maxf)))

do_print("similarity_scores.xlsx", 355)
